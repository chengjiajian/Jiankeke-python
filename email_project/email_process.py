from usages.settings import *
from usages.email_usage import pop_usage
from usages.mysql_usage import MysqlUsage,MysqlAnalyzeUsage,calculate_differ,get_all_info
from usages.file_tools import *
from email.parser import Parser
from email.header import decode_header,Header
from email.utils import parseaddr,formataddr
from xlrd import open_workbook
from openpyxl import load_workbook
from datetime import timedelta
import re
from xlrd import xldate_as_datetime
from decimal import *
import time
from dateutil.parser import parse


# 处理邮件
class ProcessEmail:
    def __init__(self):
        print('正在初始化邮箱')
        pass

    # 获取文本编码
    def guess_charset(self, msg):
        charset = msg.get_charset()
        if charset is None:
            content_type = msg.get('Content-Type', '').lower()
            pos = content_type.find('charset=')
            if pos >= 0:
                charset = content_type[pos + 8:].strip()
        return charset

    # 解码
    def decode_str(self, s):
        value, charset = decode_header(s)[0]
        if charset:
            value = value.decode(charset)
        return value

    # 获取email 附件
    def get_excel_from_email(self):
        with pop_usage() as server:
            # 使用list()返回所有邮件的编号，默认为字节类型的串
            resp, mails, octets = server.list()
            total_mail_numbers = len(mails)
            current_path = 'static'
            if not os.path.exists(main_path):  # 如果文件夹不存在，创建文件夹
                os.makedirs(main_path)
            sum_email_info = {'status': 0, 'mail_list': []}  # 记录所有 需要录的邮件信息
            # 从后往前 翻一遍邮箱
            for i in range(total_mail_numbers, 0, -1):
                resp, lines, octets = server.retr(i)
                msg_content = b'\r\n'.join(lines).decode('utf-8')
                msg = Parser().parsestr(msg_content)
                # 转换邮箱内的时间
                date_now = time.strftime("%Y%m%d-%H-%M-%S",
                                         time.strptime(msg.get("Date")[0:24], '%a, %d %b %Y %H:%M:%S'))  # 格式化收件时间
                x = time.strptime(msg.get("Date")[0:24], '%a, %d %b %Y %H:%M:%S')
                a = '{}-{}-{} {}:{}:{}'.format(x.tm_year, x.tm_mon, x.tm_mday, x.tm_hour, x.tm_min, x.tm_sec)
                delete_time = datetime.datetime.now() - datetime.timedelta(days=10)
                time_delta = parse(a) - delete_time
                if time_delta.days > 0:
                    sender = msg.get('From', '')  # 发件人信息处理========================
                    hdr, address = parseaddr(sender)
                    name = self.decode_str(hdr)
                    # 这里获得了 发件人名字，发件人邮箱
                    from_name = u'{}'.format(name)  # 发件人名称
                    from_address = u'{}'.format(address)  # 发件人邮箱
                    title_value = msg.get('Subject', '')
                    if title_value:
                        # 需要解码Subject字符串:
                        value = self.decode_str(title_value)
                        print('{}、这封邮件标题:'.format(i), value)
                        for keywords in [keywords1, keywords2, keywords3, keywords4]:
                            if keywords == value:
                                # time title是这封邮件的唯一标识 名 字
                                time_title = '{}{}'.format(keywords, date_now)
                                save_path = os.path.join(main_path, keywords, time_title)
                                if not os.path.exists(save_path):  # 如果文件夹不存在，创建文件夹
                                    os.makedirs(save_path)
                                # 检查库里情况===========================
                                check_response = MysqlUsage.check_attachment_exist(time_title)
                                if check_response:
                                    sum_email_info['status'] = 1
                                    # 遍历邮件内容
                                    for part in msg.walk():
                                        # 分析附件名字
                                        filename = part.get_filename()
                                        if filename:
                                            # 文件名转码
                                            true_file_name = self.decode_str(filename)
                                            if 'xls' in true_file_name:
                                                # 发现附件 就解析附件
                                                data = part.get_payload(decode=True)
                                                with open(os.path.join(save_path, true_file_name), 'wb') as file:
                                                    file.write(data)
                                                print('{} 下载成功'.format(true_file_name))
                                                if keywords == keywords1:
                                                    email_type = 0
                                                elif keywords == keywords2:
                                                    email_type = 1
                                                elif keywords == keywords3:
                                                    email_type = 2
                                                elif keywords == keywords4:
                                                    email_type = 3
                                                MysqlUsage.insert_data(time_title=time_title, file_name=true_file_name,
                                                                       email_type=email_type,
                                                                       email_account=from_address,
                                                                       email_name=from_name, origin_name=true_file_name)
                                                if time_title not in sum_email_info['mail_list']:
                                                    sum_email_info['mail_list'].append(time_title)
                                break
                        else:
                            message = '发现新邮件,但邮件抬头里没有台账关键字,所以跳过'
                            print(message)
                else:
                    z = server.dele(i)
        return sum_email_info


# 处理零售业务
class ProcessExcelRetail:
    def __init__(self, insurance_list, sign):
        self.insurance_list = insurance_list
        self.sign = sign
    # 读取 零售和过账不垫资 表格
    def load_excel_retail(self, excel_name):
        import re
        name_reg = re.compile('代理人-(\d{1,2})月')
        try:
            self.excel = open_workbook(excel_name)
            sheet_names = self.excel.sheet_names()
            for name in sheet_names:
                month = re.findall(name_reg, name)
                if month:
                    self.month = int(month[0])
                    break
        except Exception:
            self.excel = load_workbook(self.excel_name)
            sheet_names = self.excel.sheetnames
            print(sheet_names)
            input('')

    # 处理零售业务的表格
    def process_retail_data(self, sheet):
        company_name_dict = {'太平洋': '太平洋保险',
                                  '平安': '平安保险',
                                  '人保': '人保',
                                  '中保': '',
                                  '中华联合': '中华联合',
                                  '天安': '天安保险',
                                  '利宝': '利宝车险',
                                  '永诚': '永诚财产保险',
                                  '华安': '华安保险',
                                  '大地': '大地保险',
                                  '华泰': '华泰保险',
                                  '锦泰': '锦泰保险',
                                  '阳光': '阳光保险',
                                  '中煤': '中煤保险',
                                  '珠峰': '珠峰保险',
                                  '众安保险': '众安保险',
                                    '众安':'众安保险',
                                  '太保': '太平保险',
                                  '浙商': '浙商保险'
                                  }
        insurance_info_insert_list = []
        for i in range(sheet.nrows):
            row_value = sheet.row_values(i)
            if '保险公司' in row_value:
                print('检测到标题')
                head_row = i
                insur_sum_price_column = row_value.index('保费合计')
                insur_type_column   = row_value.index('保险公司分类')
                plate_num_column    = row_value.index('车牌号')
                owner_name_column   = row_value.index('被保险人')
                comm_price_column   = row_value.index('商业险')
                comm_rate_column    = row_value.index('商业险手续费率（收）')
                comm_income_column  = row_value.index('应收商业险手续费收入')
                comp_price_column   = row_value.index('交强险')
                comp_rate_column    = row_value.index('交强险手续费率（收）')
                comp_income_column  = row_value.index('应收交强险手续费收入')
                id_num_column       = row_value.index('身份证号码')
                try:
                    sign_date_column    = row_value.index('签单日期')
                except Exception:
                    sign_date_column = row_value.index('日期')
                comm_num_column     = row_value.index('商业保单号')
                comp_num_column     = row_value.index('交强保单号')
                insur_income_column = row_value.index('应收手续费收入合计')
                break
        insur_price = 0
        insur_income_price = 0
        # 条数统计，暂时不用
        count_times = 0
        status = 2
        uName = 'cjj_python'
        createTime = datetime.datetime.now()
        uid = 8
        flag = 1
        type_flag = 2
        cationSign = 'SYCX{}'.format(self.sign)
        print('开始处理子信息')
        for i in range(head_row + 1, sheet.nrows):
            row_value = sheet.row_values(i)
            if row_value[0]:
                count_times += 1
                #后端数据处理
                organization = '四川聚仁'
                insuranceCompany = row_value[insur_type_column]
                company = company_name_dict[insuranceCompany]
                insuranceOrgnization = '四川聚仁保险代理有限公司'
                try:
                    writeDate = xldate_as_datetime(row_value[sign_date_column],0)
                except Exception:
                    writeDate = xldate_as_datetime(row_value[sign_date_column-1],0)
                insurancePerson = row_value[owner_name_column]
                insuranceIdCard = row_value[id_num_column]
                licenseNumber = row_value[plate_num_column]
                comm_num = row_value[comm_num_column]
                comp_num = row_value[comp_num_column]
                if comm_num:
                    comm_income = Decimal(row_value[comm_income_column]).quantize(Decimal('0.00')) if row_value[comm_income_column] else Decimal('0.00')
                    comm_price = Decimal(row_value[comm_price_column]).quantize(Decimal('0.00')) if row_value[comm_price_column] else Decimal('0.00')
                    if (writeDate, comm_income, comm_num) not in self.insurance_list:
                        comm_rate  = Decimal(row_value[comm_rate_column]).quantize(Decimal('0.00')) if row_value[comm_rate_column] else Decimal(0)
                        insuranceType = 2
                        insuranceName = '商业险'
                        single_info = (organization, company, insuranceCompany, insuranceOrgnization, writeDate, insurancePerson, insuranceIdCard, licenseNumber, comm_price, comm_rate, comm_income, status, uid, uName, createTime, flag, type_flag, insuranceType, insuranceName, comm_num, self.sign, cationSign, comm_price, comm_rate, 0, 1,'','')
                        self.insurance_list.append((writeDate, comm_price, comm_num))
                        insurance_info_insert_list.append(single_info)
                if comp_num:
                    comp_price = Decimal(row_value[comp_price_column]).quantize(Decimal('0.00')) if row_value[comp_price_column] else Decimal('0.00')
                    comp_income = Decimal(row_value[comp_income_column]).quantize(Decimal('0.00')) if row_value[comp_income_column] else Decimal('0.00')
                    if (writeDate, comp_income, comp_num) not in self.insurance_list:
                        comp_rate = Decimal(row_value[comp_rate_column]).quantize(Decimal('0.00')) if row_value[comp_rate_column] else Decimal(0)
                        insuranceType = 1
                        insuranceName = '交强险'
                        single_info = (organization, company, insuranceCompany, insuranceOrgnization, writeDate, insurancePerson, insuranceIdCard, licenseNumber, comp_price, comp_rate, comp_income, status, uid, uName, createTime, flag, type_flag, insuranceType, insuranceName, comp_num, self.sign, cationSign, comp_price, comp_rate, 0, 1,'','')
                        self.insurance_list.append((writeDate, comp_price, comp_num))
                        insurance_info_insert_list.append(single_info)
                # 前端数据处理
                if row_value[insur_sum_price_column] != '':
                    insur_price += float(row_value[insur_sum_price_column])
                if row_value[insur_income_column] != '':
                    insur_income_price += float(row_value[insur_income_column])
            else:
                break
        print('子信息处理完毕')
        MysqlAnalyzeUsage.insert_single_insur(insurance_info_insert_list)
        insur_price = Decimal(round(insur_price, 2) / 10000).quantize(Decimal('0.000000'))
        insur_income_price = Decimal(round(insur_income_price, 2) / 10000).quantize(Decimal('0.000000'))
        return [insur_price, insur_income_price, count_times]  # , count_times# 条数，暂时不用

    # 处理零售业务===============================================
    def load_sheet_retail(self):
        try:
            sheet1 = self.excel.sheet_by_name('代理人-{}月'.format(self.month))
        except Exception:
            sheet1 = self.excel.sheet_by_name('代理人-0{}月'.format(self.month))
        try:
            sheet2 = self.excel.sheet_by_name('批发其他-代理点-{}月'.format(self.month))
        except Exception:
            sheet2 = self.excel.sheet_by_name('批发其他-代理点-0{}月'.format(self.month))
        print('正在处理 代理人表')
        data1 = self.process_retail_data(sheet1)
        print('正在处理 批发其他-代理点表')
        data2 = self.process_retail_data(sheet2)
        sum_data = []
        for i, value in enumerate(data1):
            sum_data.append(value + data2[i])
        return sum_data

    # 处理不垫资业务===============================================
    def load_sheet_unloan(self):
        try:
            sheet = self.excel.sheet_by_name('批发过账-{}月'.format(self.month))
        except Exception:
            sheet = self.excel.sheet_by_name('批发过账-0{}月'.format(self.month))
        for i in range(sheet.nrows):
            row_value = sheet.row_values(i)
            if '保险公司' in row_value:
                head_row = i
                tax_column = row_value.index('应收手续费收入合计')
                break
        tax_price = 0
        # 条数统计，暂时不用
        count_times = 0
        for i in range(head_row + 1, sheet.nrows):
            row_value = sheet.row_values(i)
            if row_value[1]:
                count_times += 1
                if row_value[tax_column] != '':
                    tax_price += float(row_value[tax_column])
            else:
                break
        tax_price_min = Decimal(round(tax_price, 2) / 10000).quantize(Decimal('0.000000'))
        insur_price = Decimal((tax_price * 4) / 10000).quantize(Decimal('0.000000'))
        return [insur_price, tax_price_min, count_times]  # , count_times# 条数，暂时不用


# 处理垫资业务
class ProcessExcelLoan:
    def __init__(self, insurance_list, sign):
        self.sign = sign
        self.insurance_list = insurance_list
        self.insurance_type_dict = {
            'C00': '商业险',
            'C01': '商业险',
            'C51': '交强险',
            'C50': '交强险',
            'DAA': '商业险',
            'ZDM': '诉讼财产保全责任保险',
            'YEL': '货运险',
            'YDL': '国内联运货运险',
            'ZBV': '货运险',
            'YDG': '货运险',
            'AAN': '通用航空保险',
            'EAU': '驾乘人员意外',
            'EDD': '驾驶人意外',
            'PL0700231': '职业责任险',
            'PL03J0092': '意外住院补贴',
            'PL03Y0251': '驾乘人员意外',
            'PL03Y0273': '法定节假日驾乘人员意外',
            'PL03Y0006': '驾驶人意外',
            'PL03J0003': '意外医疗',
            'PL03J0043': '意外住院收入',
            'PL03Y0111': '个人意外',
            'PL03Y0162': '交通意外',
            'PL03Y0336': '个人法律补偿',
            'GAHI': '阳光人寿附加团体意外伤害住院津贴医疗保险',
            'GADD01': '阳光人寿和泰团体意外伤害保险',
            'GAMR': '阳光人寿附加团体意外伤害医疗保险',
            '商业险': '商业险',
            '交强险': '交强险'}

    def load_excel_info(self, folder, single_excel_name):
        excel = open_workbook(os.path.join(folder,single_excel_name))
        self.sheet = excel.sheet_by_index(0)
        self.excel_name = single_excel_name

    def process_data(self):
        head_row = ''
        status = 2
        uName = 'cjj_python'
        createTime = datetime.datetime.now()
        uid = 8
        flag = 1
        type_flag = 1
        cationSign = 'SYCX{}'.format(self.sign)
        organization = '四川聚仁'
        company = '平安保险'
        insuranceCompany = '平安'
        insuranceOrgnization = '四川聚仁保险代理有限公司'
        insurance_info_insert_list = []
        loan_insur_price = 0
        loan_income = 0
        total_dict = {}
        for i in range(0, self.sheet.nrows):
            row_value = self.sheet.row_values(i)
            if '领导审批:' in row_value:
                break
            elif head_row:
                writeDate = parse(row_value[sign_date_column])
                date_format = writeDate.strftime("%Y-%m")
                if date_format not in total_dict:
                    total_dict[date_format] = {'loan_insur_price': 0,
                                               'loan_income': 0}
                insur_num = row_value[insur_num_column]
                insur_code_type = row_value[insur_type_column]
                insuranceName = self.insurance_type_dict[
                    insur_code_type] if insur_code_type in self.insurance_type_dict else insur_code_type
                if insuranceName == '商业险':
                    insuranceType = '2'
                elif insuranceName == '交强险':
                    insuranceType = '1'
                else:
                    insuranceType = '3'
                owner_name = row_value[owner_name_column]
                insur_sum_price = Decimal(row_value[insur_sum_price_column]).quantize(Decimal('0.00'))
                income = Decimal(row_value[income_column]).quantize(Decimal('0.00'))
                total_dict[date_format]['loan_insur_price'] += insur_sum_price
                total_dict[date_format]['loan_income'] += income
                rate = Decimal(row_value[rate_column]).quantize(Decimal('0.00'))
                additional_info = row_value[additional_info_column]
                if (writeDate, income, insur_num) not in self.insurance_list:
                    single_info = (
                    organization, company, insuranceCompany, insuranceOrgnization, writeDate, owner_name, '', '',insur_sum_price, rate, income, status, uid, uName, createTime, flag, type_flag, insuranceType,insuranceName, insur_num, self.sign, cationSign, income, rate, 1, 1,'','')
                    self.insurance_list.append((writeDate, income, insur_num))
                    insurance_info_insert_list.append(single_info)
            elif '支付申请号生成日期:' in row_value:
                for i in row_value:
                    try:
                        month_value = parse(i).month
                        excel_month = month_value if month_value > 9 else '0{}'.format(month_value)
                        print(excel_month)
                        break
                    except Exception:
                        pass
            elif '保单号' in row_value:
                print('检测到标题')
                head_row = i
                insur_num_column = row_value.index('保单号')
                insur_type_column = row_value.index('险种')
                owner_name_column = row_value.index('客户')
                insur_sum_price_column = row_value.index('保费收入')
                sign_date_column = row_value.index('出单时间')
                rate_column = row_value.index('本次比例%')
                income_column = row_value.index('本次手续费/经纪费(人民币)')
                additional_info_column = row_value.index('合作网点')
        MysqlAnalyzeUsage.insert_single_insur(insurance_info_insert_list)
        for month_date_format, value in total_dict.items():
            year_date, month_date = month_date_format.split('-')
            loan_insur_price = Decimal(value['loan_insur_price'] / 10000).quantize(Decimal('0.000000'))
            loan_income = Decimal(value['loan_income'] / 10000).quantize(Decimal('0.000000'))
            data_dict = {'month': month_date,
                         1: {'inPrice': loan_insur_price,
                             'inCharges': loan_income}}
            MysqlAnalyzeUsage.insert_data(data_dict, year_time=year_date,file_name=self.excel_name)


# 处理全联文件
class QuanlianExcelProcess:
    def __init__(self, insurance_list, sign):
        self.sign = sign
        self.insurance_type_dict = {
            'C00': '商业险',
            'C01': '商业险',
            'C51': '交强险',
            'C50': '交强险',
            'DAA': '商业险',
            'ZDM': '诉讼财产保全责任保险',
            'YEL': '货运险',
            'YDL': '国内联运货运险',
            'ZBV': '货运险',
            'YDG': '货运险',
            'AAN': '通用航空保险',
            'EAU': '驾乘人员意外',
            'EDD': '驾驶人意外',
            'PL0700231': '职业责任险',
            'PL03J0092': '意外住院补贴',
            'PL03Y0251': '驾乘人员意外',
            'PL03Y0273': '法定节假日驾乘人员意外',
            'PL03Y0006': '驾驶人意外',
            'PL03J0003': '意外医疗',
            'PL03J0043': '意外住院收入',
            'PL03Y0111': '个人意外',
            'PL03Y0162': '交通意外',
            'PL03Y0336': '个人法律补偿',
            'GAHI': '阳光人寿附加团体意外伤害住院津贴医疗保险',
            'GADD01': '阳光人寿和泰团体意外伤害保险',
            'GAMR': '阳光人寿附加团体意外伤害医疗保险',
            '商业险': '商业险',
            '交强险': '交强险'}
        self.insurance_list = insurance_list
        self.header_dict = {
            '合肥人保': {
                'sheet_index': 0,
                'company': '人保',
                'insuranceCompany': '合肥人保',
                'header_feature': '结算单号',
                'insur_num_name': '业务号',
                'insur_type_name': '险种',
                'owner_name_name': '投保人',
                'insur_sum_price_name': '实收保费',
                'rate_name': '基本佣金比例',
                'income_name': '应付佣金金额',
                'end_feature': '',
                'calculate_rate': 1},
            '滁州平安': {
                'sheet_index': 0,
                'company': '平安保险',
                'insuranceCompany': '滁州平安',
                'header_feature': '保单号',
                'insur_num_name': '保单号',
                'insur_type_name': '险种',
                'owner_name_name': '客户',
                'insur_sum_price_name': '保费收入',
                'rate_name': '本次比例%',
                'income_name': '本次手续费/经纪费(人民币)',
                'end_feature': '领导审批:',
                'calculate_rate': 0.99},
            '宁波阳光': {
                'sheet_index': 1,
                'company': '阳光保险',
                'insuranceCompany': '宁波阳光',
                'header_feature': '保单号',
                'insur_num_name': '保单号',
                'insur_type_name': '险种代码',
                'owner_name_name': '客户名称',
                'insur_sum_price_name': '保险费',
                'rate_name': '手续费比例',
                'income_name': '手续费',
                'end_feature': '合计',
                'calculate_rate': 0.985},
            '杭州阳光': {
                'sheet_index': 0,
                'company': '阳光保险',
                'insuranceCompany': '杭州阳光',
                'header_feature': '签单日期',
                'insur_num_name': '保单号/批单号',
                'insur_type_name': '险种',
                'owner_name_name': '被保险人',
                'insur_sum_price_name': '不含税保费',
                'rate_name': '经纪费率',
                'income_name': '经纪费金额',
                'end_feature': '',
                'calculate_rate': 0.99},
            '厦门平安': {
                'sheet_index': 0,
                'company': '平安保险',
                'insuranceCompany': '厦门平安',
                'header_feature': '保单号',
                'insur_num_name': '保单号',
                'insur_type_name': '险种',
                'owner_name_name': '客户',
                'insur_sum_price_name': '保费收入',
                'rate_name': '本次比例%',
                'income_name': '本次手续费/经纪费(人民币)',
                'end_feature': '领导审批:',
                'calculate_rate': 0.99},
            '合肥平安': {
                'sheet_index': 0,
                'company': '平安保险',
                'insuranceCompany': '合肥平安',
                'header_feature': '保单号',
                'insur_num_name': '保单号',
                'insur_type_name': '险种',
                'owner_name_name': '客户',
                'insur_sum_price_name': '保费收入',
                'rate_name': '本次比例%',
                'income_name': '本次手续费/经纪费(人民币)',
                'end_feature': '领导审批:',
                'calculate_rate': 0.99},
            '上海平安': {
                'sheet_index': 0,
                'company': '平安保险',
                'insuranceCompany': '上海平安',
                'header_feature': '保单号',
                'insur_num_name': '保单号',
                'insur_type_name': '险种',
                'owner_name_name': '客户',
                'insur_sum_price_name': '保费收入',
                'rate_name': '本次比例%',
                'income_name': '本次手续费/经纪费(人民币)',
                'end_feature': '领导审批:',
                'calculate_rate': 0.99},
            '平安非垫资': {
                'sheet_index': 0,
                'company': '平安保险',
                'insuranceCompany': '平安保险',
                'header_feature': '保单号',
                'insur_num_name': '保单号',
                'insur_type_name': '险种',
                'owner_name_name': '客户',
                'insur_sum_price_name': '保费收入',
                'rate_name': '本次比例%',
                'income_name': '本次手续费/经纪费(人民币)',
                'end_feature': '领导审批:',
                'calculate_rate': 0.99},
        }

    def process_data(self, folder, single_excel_name):
        print('正在检测文件：{}'.format(single_excel_name))
        date_info = re.findall(re.compile('20\d{6}'), single_excel_name)
        if date_info:
            date_info = date_info[0]
            file_path = os.path.join(folder, single_excel_name)
            uName = 'cjj_python'
            uid = 8
            flag = 1
            status = 2
            type_flag = 1
            createTime = datetime.datetime.now()
            organization = '全联'
            insuranceOrgnization = '全联保险经纪有限公司'
            insurance_info_insert_list = []
            total_dict = {}
            year_date = date_info[:4]
            month_date = date_info[4:6]
            day_date = date_info[-2:]
            head_row = ''
            reg = re.compile('_([A-Z]{2}-201[0-9]{5}-[0-9]{3})')
            contractId = re.findall(reg, single_excel_name)[0]
            proName = MysqlAnalyzeUsage.get_proName_by_contractId(contractId)['proName']
            # 垫资 逻辑处理
            if '非垫资' not in single_excel_name:
                data_type = 1
                print(single_excel_name)
                for insurance_company_name, value in self.header_dict.items():
                    if insurance_company_name in single_excel_name:
                        cationSign = 'SYCX{}'.format(self.sign)
                        #ratio_info_dict = MysqlAnalyzeUsage.get_ratio_info(company_name=insurance_company_name)
                        payRatio, incomeRatio, noCarRatio = (0.925 * 0.998, 0.99, 0.985)
                        # if ratio_info_dict:
                        #     ratio_info_dict = ratio_info_dict[0]
                        #     payRatio, incomeRatio, noCarRatio = ratio_info_dict['payRatio'], ratio_info_dict[
                        #         'incomeRatio'], ratio_info_dict['noCarRatio']
                        company = value['company']
                        insuranceCompany = value['insuranceCompany']
                        if re.findall('xls$', single_excel_name):
                            excel = open_workbook(file_path)
                            sheet = excel.sheet_by_index(0)
                            sheet_list = [sheet.row_values(i) for i in range(sheet.nrows)]
                        elif re.findall('xlsx$', single_excel_name):
                            excel = load_workbook(file_path, data_only=True)
                            sheetNames = excel.sheetnames
                            sheet = excel[sheetNames[0]]
                            sheet_list = [[cell.value for cell in i] for i in sheet.iter_rows()]
                        policyCount = 0
                        amount = 0
                        advance = 0
                        for row_value in sheet_list:
                            if row_value[0] == value['end_feature'] or row_value[0] in ['合计:', '合计']:
                                print('{} 已到最后一行'.format(insurance_company_name))
                                break
                            elif head_row:
                                writeDate = parse('{}/{}/{}'.format(year_date, month_date, day_date))
                                date_format = writeDate.strftime("%Y-%m")
                                if date_format not in total_dict:
                                    total_dict[date_format] = {'loan_insur_price': 0,
                                                               'loan_income': 0}
                                insur_num = row_value[insur_num_column]
                                insur_code_type = row_value[insur_type_column]
                                if insur_code_type in self.insurance_type_dict:
                                    insuranceName = self.insurance_type_dict[insur_code_type]
                                    if insuranceName == '商业险':
                                        insuranceType = '2'
                                    elif insuranceName == '交强险':
                                        insuranceType = '1'
                                    else:
                                        insuranceType = '3'
                                else:
                                    insuranceType = '3'
                                    insuranceName = insur_code_type
                                owner_name = row_value[owner_name_column]
                                insur_sum_price = Decimal(row_value[insur_sum_price_column]).quantize(Decimal('0.00'))
                                # 判断 非车险，区分非车险 车险的乘法系数
                                plus_rate = incomeRatio if insuranceType != '3' else noCarRatio
                                income = (Decimal(row_value[income_column]) * Decimal(plus_rate)).quantize(Decimal('0.00'))
                                # 收入总额 amount
                                amount += income
                                # 付款金额（原垫款金额）
                                advance += income * Decimal(payRatio)
                                total_dict[date_format]['loan_insur_price'] += insur_sum_price
                                total_dict[date_format]['loan_income'] += income
                                rate = Decimal(row_value[rate_column]).quantize(Decimal('0.00'))
                                policyCount += 1
                                if (writeDate, income, insur_num) not in self.insurance_list:
                                    single_info = (
                                        organization, company, insuranceCompany, insuranceOrgnization,
                                        writeDate, owner_name, '', '',
                                        insur_sum_price, rate, income, status, uid, uName, createTime, flag,
                                        type_flag,
                                        insuranceType,
                                        insuranceName, insur_num, self.sign, cationSign, income, plus_rate, 1, 1, contractId, proName)
                                    self.insurance_list.append((writeDate, income, insur_num))
                                    insurance_info_insert_list.append(single_info)
                            elif value['header_feature'] in row_value:
                                insur_num_column = row_value.index(value['insur_num_name'])
                                insur_type_column = row_value.index(value['insur_type_name'])
                                owner_name_column = row_value.index(value['owner_name_name'])
                                insur_sum_price_column = row_value.index(value['insur_sum_price_name'])
                                rate_column = row_value.index(value['rate_name'])
                                income_column = row_value.index(value['income_name'])
                                head_row = 1
                        cationDate = datetime.datetime.now().strftime('%Y-%m-%d')
                        advance_insert_list = [(cationDate, organization, company, cationSign, amount, policyCount,
                                                advance, 4, contractId, proName), ]
                        MysqlAnalyzeUsage.insert_loan_info(advance_insert_list)
                        break
            # 非 垫资 众安
            elif '非垫资' in single_excel_name and '众安' in single_excel_name:
                print('众安保险')
                cationSign = 'SYCX{}'.format(self.sign)
                data_type = 2
                company = '众安保险'
                insuranceCompany = '众安保险'
                if re.findall('xls$', single_excel_name):
                    excel = open_workbook(file_path)
                    sheet1, sheet2 = excel.sheet_by_index(0), excel.sheet_by_index(1)
                    sheet_list = [sheet1.row_values(i) for i in range(sheet1.nrows)]
                    true_income_value = float(sheet2.row_values(1)[0])
                elif re.findall('xlsx$', single_excel_name):
                    excel = load_workbook(file_path, data_only=True)
                    sheet1, sheet2 = excel[excel.sheetnames[0]], excel[excel.sheetnames[1]]
                    sheet_list = [[cell.value for cell in i] for i in sheet1.iter_rows()]
                    true_income_value = float(sheet2.cell(row=2, column=1).value)
                # true_income_value = float(sheet2.cell(row=2, column=1).value)
                # rows = len(sheet_list)
                amount = true_income_value
                advance = amount * 0.93
                policyCount = 0
                print(true_income_value)
                for row_value in sheet_list:
                    # row_value = sheet_list[i]
                    if row_value[0] == '':
                        print('{} 已到最后一行'.format(single_excel_name))
                        break
                    elif head_row:
                        writeDate = parse('{}/{}/{}'.format(year_date, month_date, day_date))
                        date_format = writeDate.strftime("%Y-%m")
                        if date_format not in total_dict:
                            total_dict[date_format] = {'loan_insur_price': 0,
                                                       'loan_income': 0}
                        insur_num = row_value[insur_num_column]
                        insur_code_type = row_value[insur_type_column]
                        total_dict[date_format]['loan_income'] = true_income_value
                        if insur_code_type in self.insurance_type_dict:
                            insuranceName = self.insurance_type_dict[insur_code_type]
                            if insuranceName == '商业险':
                                insuranceType = '2'
                            elif insuranceName == '交强险':
                                insuranceType = '1'
                            else:
                                insuranceType = '3'
                        else:
                            insuranceType = '3'
                            insuranceName = insur_code_type
                        owner_name = row_value[owner_name_column]
                        comm_price = row_value[comm_price_column] if row_value[comm_price_column] else 0
                        comp_price = row_value[comp_price_column] if row_value[comp_price_column] else 0
                        income_price = row_value[income_column] if row_value[income_column] else 0
                        insur_sum_price = Decimal(comm_price + comp_price).quantize(Decimal('0.00'))
                        income = Decimal(income_price).quantize(Decimal('0.00'))
                        total_dict[date_format]['loan_insur_price'] += insur_sum_price
                        # total_dict[date_format]['loan_income'] += income   众安无法按照单条累积 计算总的佣金，先注释预留
                        policyCount += 1
                        rate = Decimal(income / insur_sum_price).quantize(
                            Decimal('0.00')) if insur_sum_price != 0 else 0
                        if (writeDate, income, insur_num) not in self.insurance_list:
                            single_info = (
                                organization, company, insuranceCompany, insuranceOrgnization,
                                writeDate, owner_name,
                                '', '',
                                insur_sum_price, rate, income, status, uid, uName, createTime, flag,
                                type_flag,
                                insuranceType,
                                insuranceName, insur_num, self.sign, cationSign, 0, 0, 2, 4, contractId, proName)
                            self.insurance_list.append((writeDate, income, insur_num))
                            insurance_info_insert_list.append(single_info)
                    elif '保单号' in row_value:
                        insur_num_column = row_value.index('保单号')
                        insur_type_column = row_value.index('险种')
                        owner_name_column = row_value.index('投保人')
                        comm_price_column = row_value.index('商业险签单保费（元）')
                        comp_price_column = row_value.index('交强险保费（元）')
                        # insur_sum_price_column = row_value.index(value['insur_sum_price_name'])
                        # rate_column = row_value.index(value['rate_name'])
                        income_column = row_value.index('费用（元）')
                        head_row = 1
                cationDate = datetime.datetime.now().strftime('%Y-%m-%d')
                advance_insert_list = [(cationDate, organization, company, cationSign, amount, policyCount,
                                        advance, 4, contractId, proName), ]
                MysqlAnalyzeUsage.insert_loan_info(advance_insert_list)
            # 非 垫资 平安
            elif '非垫资' in single_excel_name:
                data_type = 2
                for insurance_company_name, value in self.header_dict.items():
                    if insurance_company_name in single_excel_name:
                        cationSign = 'SYCX{}'.format(self.sign)
                        ratio_info_dict = MysqlAnalyzeUsage.get_ratio_info(company_name=insurance_company_name)
                        payRatio, incomeRatio, noCarRatio = (0.925, 1, 1)
                        if ratio_info_dict:
                            ratio_info_dict = ratio_info_dict[0]
                            incomeRatio, noCarRatio = ratio_info_dict['incomeRatio'], ratio_info_dict['noCarRatio']
                        company = value['company']
                        insuranceCompany = value['insuranceCompany']
                        if re.findall('xls$', single_excel_name):
                            excel = open_workbook(file_path)
                            sheet = excel.sheet_by_index(0)
                            sheet_list = [sheet.row_values(i) for i in range(sheet.nrows)]
                        elif re.findall('xlsx$', single_excel_name):
                            excel = load_workbook(file_path, data_only=True)
                            sheetNames = excel.sheetnames
                            sheet = excel[sheetNames[1]] if '中国银行' in sheetNames else excel[sheetNames[0]]
                            sheet_list = [[cell.value for cell in i] for i in sheet.iter_rows()]
                        policyCount = 0
                        amount = 0  # 收入金额
                        advance = 0  # 付款金额
                        for row_value in sheet_list:
                            if row_value[0] in [value['end_feature'], '合计:']:
                                print('{} 已到最后一行'.format(insurance_company_name))
                                break
                            elif head_row:
                                writeDate = parse('{}/{}/{}'.format(year_date, month_date, day_date))
                                date_format = writeDate.strftime("%Y-%m")
                                if date_format not in total_dict:
                                    total_dict[date_format] = {'loan_insur_price': 0,
                                                               'loan_income': 0}
                                insur_num = row_value[insur_num_column]
                                insur_code_type = row_value[insur_type_column]
                                if insur_code_type in self.insurance_type_dict:
                                    insuranceName = self.insurance_type_dict[insur_code_type]
                                    if insuranceName == '商业险':
                                        insuranceType = '2'
                                    elif insuranceName == '交强险':
                                        insuranceType = '1'
                                    else:
                                        insuranceType = '3'
                                else:
                                    insuranceType = '3'
                                    insuranceName = insur_code_type
                                owner_name = row_value[owner_name_column]
                                insur_sum_price = Decimal(row_value[insur_sum_price_column]).quantize(
                                    Decimal('0.00'))
                                # 判断 非车险，区分非车险 车险的乘法系数
                                plus_rate = incomeRatio if insuranceType != '3' else noCarRatio
                                income = (Decimal(row_value[income_column]) * Decimal(plus_rate)).quantize(Decimal('0.00'))
                                # 收入
                                amount += income
                                # 垫款金额
                                advance += income * Decimal(payRatio)
                                total_dict[date_format]['loan_insur_price'] += insur_sum_price
                                total_dict[date_format]['loan_income'] += income
                                rate = Decimal(row_value[rate_column]).quantize(Decimal('0.00'))
                                policyCount += 1
                                if (writeDate, income, insur_num) not in self.insurance_list:
                                    single_info = (
                                        organization, company, insuranceCompany, insuranceOrgnization,
                                        writeDate, owner_name,
                                        '', '',
                                        insur_sum_price, rate, income, status, uid, uName, createTime, flag,
                                        type_flag,
                                        insuranceType,
                                        insuranceName, insur_num, self.sign, cationSign, 0, 0, 2, 4, contractId, proName)
                                    self.insurance_list.append((writeDate, income, insur_num))
                                    insurance_info_insert_list.append(single_info)
                            elif value['header_feature'] in row_value:
                                insur_num_column = row_value.index(value['insur_num_name'])
                                insur_type_column = row_value.index(value['insur_type_name'])
                                owner_name_column = row_value.index(value['owner_name_name'])
                                insur_sum_price_column = row_value.index(value['insur_sum_price_name'])
                                rate_column = row_value.index(value['rate_name'])
                                income_column = row_value.index(
                                    '合计支付含增值税金额') if '合计支付含增值税金额' in row_value else row_value.index(
                                    value['income_name'])
                                head_row = 1
                        cationDate = datetime.datetime.now().strftime('%Y-%m-%d')
                        advance_insert_list = [(cationDate, organization, company, cationSign, amount, policyCount,
                                                advance, 4, contractId, proName), ]
                        MysqlAnalyzeUsage.insert_loan_info(advance_insert_list)
                        break
            if total_dict:
                for month_date_format, value in total_dict.items():
                    print(month_date_format, value)
                    year_date, month_date = month_date_format.split('-')
                    loan_insur_price = Decimal(value['loan_insur_price'] / 10000).quantize(Decimal('0.000000'))
                    loan_income = Decimal(value['loan_income'] / 10000).quantize(Decimal('0.000000'))
                    data_dict = {'month': month_date,
                                 data_type: {'inPrice': loan_insur_price,
                                             'inCharges': loan_income}}
                    MysqlAnalyzeUsage.insert_data(data_dict, organization_name='狮域全联', year_time=year_date, file_name=single_excel_name)
            # 插 单条库操作
            MysqlAnalyzeUsage.insert_single_insur(insurance_info_insert_list)
        else:
            print('文件名日期格式出错')


# 处理批增业务
class processAdditionalMoney():
    def __init__(self):
        import collections
        self.header_dict = collections.OrderedDict()
        self.header_dict['手续费(CNY)'] = {
            'company': '国寿保险',
            'insuranceCompany': '国寿',
            'organization': '师域光华',
            'ownerName': '投保人',
            'returnPrice': "手续费(CNY)",
            'mark': ['险种名称', ],
            "price_offset": 0,
            'end_feature_list': ['合计：', '合计:', '合计', '', None]}
        self.header_dict['保批单号'] = {
            'company': '阳光保险',
            'insuranceCompany': '阳光',
            'organization': '师域',
            'ownerName': '车牌号',
            'returnPrice': "金额",
            'mark': ['保批单号', ],
            "price_offset": 0,
            'end_feature_list': ['合计：', '总计:', '合计', '', None]}
        self.header_dict['投放金额'] = {#车主	车架号	手机号	投放项目	投放时间	投放金额，
                'company': '平安保险',
                'insuranceCompany': '平安',
                'organization': 'sy_dt',
                'ownerName': '车主',
                'returnPrice': "投放金额",
                'mark': ['投放', '投放项目'],
                'price_offset': 0,
                'end_feature_list': ['', None]}
        self.header_dict['佣金跟单比例'] = {#序号	结算单号	业务号	归属机构	交叉销售机构	险种	币别	实收保费	佣金跟单比例			佣金	基本佣金比例	基本佣金	职级佣金比例	职级佣金	应付佣金金额	以往已支付比例	本次应支付比例	起保日期	终保日期
                'company': '人民保险',
                'insuranceCompany': '人保',
                'organization': '全联',
                'ownerName': '投保人',
                'returnPrice': "佣金",
                'mark': ['归属机构', ],
                "price_offset": -1,
                'end_feature_list': ['合计:', '合计', '', None]}
        self.header_dict['单价/小时'] = {
            'company': '平安保险',
            'insuranceCompany': '平安',
            'organization': '狮域',
            'ownerName': '姓名',
            'returnPrice': "金额",
            'mark': ['服务名称', ],
            "price_offset": 0,
            'end_feature_list': ['合计:', '合计', '', None]}
        self.header_dict['金额'] = {
            'company': '国寿保险',
            'insuranceCompany': '国寿',
            'organization': '狮域',
            'ownerName': '序号',
            'returnPrice': "金额",
            'mark': ['车架号', ],
            "price_offset": 0,
            'end_feature_list': ['合计:', '合计', '', None]}


    def processExcel(self, sheet_list, cationNumber, states=5):
        # 常量
        row_count = 0
        amount_count = 0
        uName = 'cjj_python_additional'
        uid = 8
        flag = 1
        status = 2
        type_flag = 1
        nesstype = 2
        postpone = 2
        head_row = ''
        insert_list = []  # 录入列表

        # 预设量
        ownerName_column = 0
        returnPrice_column = 0
        mark_column = 0
        price_offset = 0
        end_feature_list = ['', ]
        for row_value in sheet_list:
            print(row_value)
            # 监测是否到最后一行
            if row_value[0] in end_feature_list:
                break
            # 标题已确定
            elif head_row:
                ownerName = row_value[ownerName_column]
                #print(returnPrice_column + price_offset)
                returnPrice = round(float(row_value[returnPrice_column + price_offset]), 2)
                row_count += 1
                amount_count += returnPrice
                mark = "" if mark_column == 0 else row_value[mark_column]
                detail_list = [company, insuranceCompany, organization, ownerName, returnPrice, mark, uName, uid, flag, status, type_flag, nesstype, states, postpone, cationNumber]  # 详细信息
                insert_list.append(detail_list)  # 插表 的 列表
            # 标题还未确定时
            elif not head_row:
                for head, value in self.header_dict.items():
                    if head in row_value:
                        company = value['company']
                        insuranceCompany = value['insuranceCompany']
                        organization = value['organization']
                        ownerName_column = row_value.index(value['ownerName'])
                        returnPrice_column = row_value.index(value['returnPrice'])
                        price_offset = value['price_offset']
                        end_feature_list = value['end_feature_list']
                        for mark in value['mark']:
                            if mark in row_value:
                                mark_column = row_value.index(mark)
                                break
                        head_row = 1
                        break
        #print(insert_list)
        MysqlAnalyzeUsage.insert_additional_detail(insert_list)
        return (row_count,amount_count)

    def getExcel(self, excel_path):
        pure_file_name, excel_type = os.path.splitext(os.path.split(excel_path)[1])
        cationNumber = re.findall('SYCX[0-9]{13}', pure_file_name)
        money_count = 0
        promt_count = 0
        if cationNumber:
            cationNumber = cationNumber[0]
            states = MysqlAnalyzeUsage.get_state_from_advance(cationNumber)
            if excel_type == '.xls':
                excel = open_workbook(excel_path)
                sheetNames = excel.sheet_names()
                for sheetName in sheetNames:
                    sheet = excel.sheet_by_name(sheetName)
                    sheet_list = [sheet.row_values(i) for i in range(sheet.nrows)]
                    row_count, amount_count = self.processExcel(sheet_list, cationNumber, states)
                    money_count += amount_count
                    promt_count += row_count
            elif excel_type == '.xlsx':
                excel = load_workbook(excel_path, data_only=True)
                sheetNames = excel.sheetnames
                for sheetName in sheetNames:
                    sheet = excel[sheetName]
                    sheet_list = [[cell.value for cell in i] for i in sheet.iter_rows()]
                    row_count, amount_count = self.processExcel(sheet_list, cationNumber, states)
                    money_count += amount_count
                    promt_count += row_count
            MysqlAnalyzeUsage.update_advance_counts(promt_count,money_count,cationNumber)
            return True
        else:
            return False


# 主任务流程
def main_job():
# try:
    write_down_error('开始执行主任务')
    ruuning_time_count = 0
    sum_list = get_all_info()
    while True:
        mail = ProcessEmail()
        email_info = mail.get_excel_from_email()
        print(email_info)
        # print(sum_list)
        if email_info['status'] == 1:
            for email_title in email_info['mail_list']:
                # 切片取 email的 前四位 来区分业务情况
                main_folder_name = email_title[:4]
                excel_folder_path = os.path.join(main_path, main_folder_name, email_title)
                sign = str(int(time.time() * 1000))
                for file in os.listdir(excel_folder_path):
                # try:
                    excel_path = os.path.join(excel_folder_path, file)
                    # print(excel_path)
                    # print(main_folder_name)
                    # print(file)
                    # input('')
                    # 婉怡：零售 + 过账不垫资业务
                    if main_folder_name == '聚仁台账':
                        # 初始化读取零售过账业务的模板
                        excel_process = ProcessExcelRetail(sum_list,sign)
                        # 模板读表信息
                        excel_process.load_excel_retail(excel_path)
                        retail_data_inPrice, retail_data_inCharge, retail_data_count = excel_process.load_sheet_retail()
                        unloan_data_inPrice, unloan_data_inCharge, unloan_data_count = excel_process.load_sheet_unloan()
                        data_info = {0: {'inPrice': retail_data_inPrice,
                                         'inCharges': retail_data_inCharge,
                                         'count': retail_data_count},
                                     2: {'inPrice': unloan_data_inPrice,
                                         'inCharges': unloan_data_inCharge,
                                         'count': unloan_data_count}}
                        print('excel data:', data_info)
                        # 因为是累加式的数据，所以我们需要做个减法在录入
                        differ = calculate_differ(data_info, month_time=excel_process.month)
                        print('after data:', differ)
                        if int((datetime.datetime(year=int(email_title[4:8]), month=excel_process.month,
                                                  day=1) - datetime.datetime.now()).days) > 0:
                            now_year = int(email_title[4:8]) - 1
                        else:
                            now_year = email_title[4:8]
                        # 插入数据
                        MysqlAnalyzeUsage.insert_data(differ, year_time=now_year, file_name=file)
                    # 婉怡：垫资业务
                    elif main_folder_name == '聚仁垫资':
                        # 初始化一个读取垫资的模板
                        excel_process = ProcessExcelLoan(sum_list,sign)
                        # 模板读表
                        excel_process.load_excel_info(folder=excel_folder_path, single_excel_name=file)
                        # 处理数据
                        excel_process.process_data()
                    # 文才：垫资 + 非垫资 业务
                    elif main_folder_name == '全联表格':
                        # 初始化一个读取全联的模板
                        excel_process = QuanlianExcelProcess(sum_list, sign)
                        excel_process.process_data(folder=excel_folder_path, single_excel_name=file)
                    elif main_folder_name == '批增表格':
                        excel_process = processAdditionalMoney()
                        excel_process.getExcel(excel_path)
                # except Exception as e:
                #     if '聚仁' in main_folder_name:
                #         sendsSMS(phone=juren_phone)
                #     elif '全联' in main_folder_name or '批增' in main_folder_name:
                #         sendsSMS(phone=quanlian_phone)
                #     sendsSMS()
                #     print('发生错误{}--{}'.format(e,file))
                #     write_down_error(str(e) + file)
                MysqlAnalyzeUsage.process_middle_to_advance()
        ruuning_time_count += 1
        next_time = '本次扫描结束，下次扫描时间 {}'.format(datetime.datetime.now() + timedelta(minutes=7))
        print(next_time)
        write_down_time(next_time)
        time.sleep(60 * 7)
# except Exception as e:
#     print('发生未知错误，详情请看日志')
#     sendsSMS()
#     write_down_error('run {} times '.format(ruuning_time_count) + '\n' + str(e))
#     print('系统将在 30 分钟后重启： {}'.format(datetime.datetime.now() + timedelta(minutes=30)))
#     time.sleep(60 * 30)
#     main_job()

if __name__ == '__main__':
    main_job()